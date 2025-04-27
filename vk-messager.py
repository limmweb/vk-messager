import os
import json
import unicodedata
import vk_api
import configparser
import re
import requests
from datetime import datetime
from openai import OpenAI
import time
from requests.exceptions import ConnectionError, ReadTimeout
import threading
import openpyxl
from openpyxl.utils import get_column_letter

# Настройка директорий
SESSIONS_DIR = os.path.join(os.getcwd(), "Sessions")
DOSSIER_DIR = os.path.join(os.getcwd(), "Dossier")
REPORTS_FILE = "reports.xlsx"
if not os.path.exists(SESSIONS_DIR):
    os.makedirs(SESSIONS_DIR)
if not os.path.exists(DOSSIER_DIR):
    os.makedirs(DOSSIER_DIR)

# Функция с логикой повторного запроса (для LongPoll и других HTTP-запросов)
def retry_request(request_func, backoff_factor=1, timeout=30):
    retries = 0
    while True:
        try:
            return request_func(timeout=timeout)
        except (ConnectionError, ReadTimeout) as e:
            print(f"Ошибка соединения: {e}")
            retries += 1
            wait_time = min(backoff_factor * (2 ** retries), 60)
            print(f"Попытка {retries}. Ожидание {wait_time} секунд перед повтором.")
            time.sleep(wait_time)

# Функция для повторных попыток VK API запросов (например, отправки сообщений)
def retry_vk_request(request_func, max_retries=3, backoff_factor=1):
    retries = 0
    while retries < max_retries:
        try:
            return request_func()
        except (ConnectionError, ReadTimeout, vk_api.exceptions.ApiError) as e:
            retries += 1
            if retries == max_retries:
                raise Exception(f"Не удалось выполнить запрос после {max_retries} попыток: {e}")
            wait_time = backoff_factor * (2 ** retries)
            print(f"Ошибка VK API: {e}. Повторная попытка {retries}/{max_retries}. Ожидание {wait_time} секунд.")
            time.sleep(wait_time)

# Сканирование и выбор сессии
def scan_sessions():
    sessions = [f for f in os.listdir(SESSIONS_DIR) if f.endswith(".ini")]
    print("Выберите сессию или создайте новую:")
    for idx, session_file in enumerate(sessions, start=1):
        print(f"{idx}: {session_file}")
    print("0: создать новую сессию")
    try:
        choice = int(input("Введите номер сессии: "))
    except ValueError:
        print("Пожалуйста, введите число.")
        return scan_sessions()
    if choice == 0:
        return create_new_session()
    elif 0 < choice <= len(sessions):
        return os.path.join(SESSIONS_DIR, sessions[choice - 1])
    else:
        print("Неверный выбор.")
        return scan_sessions()

# Очистка имени файла от запрещённых символов
def clean_filename(filename):
    forbidden_chars = r'[\\/:*?"<>|]'
    cleaned = re.sub(forbidden_chars, "", filename).strip()
    return cleaned[:100]

# Создание новой сессии
def create_new_session():
    vk_token = input("Введите ваш API ключ VK: ")
    openai_token = input("Введите ваш API ключ OpenAI: ")
    is_group = input("Это сессия для группы? (да/нет): ").lower() == "да"
    group_id = None
    if is_group:
        group_id = input("Введите ID группы (например, 12345678): ")
    personality = input("Какая будет ваша личность? (до 1000 символов): ")[:1000]
    commercial_info = input("Информация о вашем бизнесе и товарах? (до 1000 символов): ")[:1000]
    conversation_rules = input("Какие будут правила общения? (до 1000 символов): ")[:1000]
    conversation_goal = input("Какие будут цели общения? (до 1000 символов): ")[:1000]

    user_info = authenticate_vk(vk_token, group_id)
    if not user_info:
        return None

    if group_id:
        vk_session = vk_api.VkApi(token=vk_token)
        vk = vk_session.get_api()
        group_info = vk.groups.getById(group_id=group_id, fields="description")[0]
        group_name = clean_filename(group_info["name"])
        group_description = group_info.get("description", "")[:1000]
        session_file = os.path.join(SESSIONS_DIR, f"group_{group_id}_{group_name}.ini")
    else:
        first_name = user_info.get("first_name", "Unknown")
        last_name = user_info.get("last_name", "Unknown")
        session_file = os.path.join(SESSIONS_DIR, f"{first_name}_{last_name}.ini")
        group_name = ""
        group_description = ""

    config = configparser.ConfigParser()
    config["DEFAULT"] = {
        "vk-token": vk_token,
        "openai-token": openai_token,
        "group-id": group_id if group_id else "",
        "group-name": group_name,
        "group-description": group_description,
        "personality": personality,
        "commercial-info": commercial_info,
        "conversation-rules": conversation_rules,
        "conversation-goal": conversation_goal,
        "tokens_in": "0",
        "tokens_out": "0",
        "tokens_total": "0",
        "tokens_cost": "0",
        **(user_info if not group_id else {"group_id": group_id}),
        "Characteristic": ""
    }

    with open(session_file, "w", encoding="utf-8") as configfile:
        config.write(configfile)
    return session_file

# Аутентификация пользователя или группы
def authenticate_vk(vk_token, group_id=None):
    try:
        vk_session = vk_api.VkApi(token=vk_token)
        vk = vk_session.get_api()
        if group_id:
            group_info = vk.groups.getById(group_id=group_id)[0]
            print(f"Успешная аутентификация! Группа: {group_info['name']} (ID: {group_id})")
            return {"group_id": group_id, "name": group_info["name"]}
        else:
            fields = "first_name,last_name"
            user_info = vk.users.get(fields=fields)[0]
            print(f"Успешная аутентификация! Пользователь: {user_info['first_name']} {user_info['last_name']}")
            return user_info
    except Exception as e:
        print(f"Ошибка аутентификации: {e}")
        return None

# Получение информации о собеседнике и обновление досье
def get_conversation_partner_info(vk, user_id):
    fields = (
        "activities, about, blacklisted, blacklisted_by_me, books, bdate, can_write_private_message, "
        "career, city, contacts, education, followers_count, friend_status, home_town, interests, last_seen, movies, music, status"
    )
    partner_info = retry_request(lambda timeout: vk.users.get(user_ids=user_id, fields=fields)[0])
    if partner_info and (partner_info.get("blacklisted") == 1 or partner_info.get("blacklisted_by_me") == 1 or partner_info.get("can_write_private_message") == 0):
        print("Собеседник заблокирован или не может писать в личные сообщения.")
        return None

    filtered_info = {k: (v[:1000] if isinstance(v, str) and len(v) > 1000 else v) for k, v in partner_info.items() if v and k not in ["blacklisted", "blacklisted_by_me", "can_write_private_message"]}
    dossier_file = os.path.join(DOSSIER_DIR, f"{partner_info['first_name']}_{partner_info['last_name']}_{user_id}.json")

    default_dossier = {
        "tokens_in": 0,
        "tokens_out": 0,
        "tokens_total": 0,
        "tokens_cost": 0,
        "photo_description": "",
        "characteristic": "",
        "sale_status": "",
        "profit": 0,
        "api_token": "",
        "token_added": "",
        "token_status": "",
        "contacts": ""
    }

    if os.path.exists(dossier_file):
        with open(dossier_file, "r", encoding="utf-8") as file:
            old_data = json.load(file)
        for key, value in filtered_info.items():
            old_data[key] = value
        for key, value in default_dossier.items():
            if key not in old_data:
                old_data[key] = value
        dossier_data = old_data
    else:
        dossier_data = {**filtered_info, **default_dossier}

    with open(dossier_file, "w", encoding="utf-8") as file:
        json.dump(dossier_data, file, ensure_ascii=False)
    return json.dumps(filtered_info, ensure_ascii=False)

# Функция для экранирования только кавычек и слэшей
def escape_json_string(text):
    return text.replace('"', '\\"').replace('\\', '\\\\')

# Функция для декодирования Unicode-последовательностей
def decode_unicode(text):
    return unicodedata.normalize('NFKD', text.encode().decode('unicode_escape'))

# Формирование запроса OpenAI
def create_openai_prompt(config, conversation_history, partner_info, entity_id):
    date_time_str = datetime.now().strftime("%A %d %B %Y, %H:%M:%S")
    personality = decode_unicode(config['DEFAULT']['personality'])
    commercial_info = decode_unicode(config['DEFAULT']['commercial-info'])
    conversation_rules = decode_unicode(config['DEFAULT']['conversation-rules'])
    conversation_goal = decode_unicode(config['DEFAULT']['conversation-goal'])
    
    system_content = (
        f"{date_time_str}\n"
        "Ты переписываешься в личных сообщениях Вконтакте от Моего лица используя указанные данные и инструкции.\n"
        f"Моя Личность: \"{personality}\"\n"
        f"Коммерческая информация: \"{commercial_info}\"\n"
        f"Мой Собеседник: \"{partner_info}\" . Информация предоставлена из API метода user.get ВКонтакте . "
        "Пожалуйста расшифровывай пары \"ключ\": \"значение\" в соответствии с известной тебе документацией.\n"
        f"Инструкции общения: \"{conversation_rules}\" , Цели общения: \"{conversation_goal}\" "
    )
    prompt = [{"role": "system", "content": system_content}]

    conversation_history = sorted(conversation_history, key=lambda msg: msg['date'])
    for msg in conversation_history:
        if not msg['text']:
            continue
        role = "assistant" if msg['from_id'] == int(entity_id) else "user"
        content = escape_json_string(decode_unicode(msg['text']))[:1000]
        prompt.append({"role": role, "content": content})

    prompt.append({"role": "system", "content": "Предоставь ответ для диалога. В ответ только готовое к отправке сообщение и ничего больше."})
    print("Цепочка сообщений для OpenAI:", prompt)
    return prompt

# Обновление токенов в сессии
def update_session_tokens(session_file, tokens):
    config = configparser.ConfigParser()
    config.read(session_file, encoding="utf-8")
    current_in = int(config["DEFAULT"].get("tokens_in", 0))
    current_out = int(config["DEFAULT"].get("tokens_out", 0))
    current_total = int(config["DEFAULT"].get("tokens_total", 0))
    current_cost = float(config["DEFAULT"].get("tokens_cost", 0))

    config["DEFAULT"]["tokens_in"] = str(current_in + tokens["input"])
    config["DEFAULT"]["tokens_out"] = str(current_out + tokens["output"])
    config["DEFAULT"]["tokens_total"] = str(current_total + tokens["total"])
    config["DEFAULT"]["tokens_cost"] = str(current_cost + tokens["cost"])

    with open(session_file, "w", encoding="utf-8") as configfile:
        config.write(configfile)

# Обновление токенов в досье
def update_dossier_tokens(dossier_file, tokens):
    with open(dossier_file, "r", encoding="utf-8") as file:
        dossier_data = json.load(file)

    dossier_data["tokens_in"] = dossier_data.get("tokens_in", 0) + tokens["input"]
    dossier_data["tokens_out"] = dossier_data.get("tokens_out", 0) + tokens["output"]
    dossier_data["tokens_total"] = dossier_data.get("tokens_total", 0) + tokens["total"]
    dossier_data["tokens_cost"] = dossier_data.get("tokens_cost", 0) + tokens["cost"]

    with open(dossier_file, "w", encoding="utf-8") as file:
        json.dump(dossier_data, file, ensure_ascii=False)

# Запись в отчёт в формате .xlsx
def log_report(account_id, account_name, message, recipient_id, recipient_name, tokens):
    timestamp = datetime.now().strftime("%A %d %B %Y, %H:%M:%S")
    entity_type = "group" if str(account_id).startswith("-") else "user"

    message = message.replace("\n", " ").replace("\r", " ")

    headers = ["timestamp", "account_id", "account_name", "entity_type", "message", "recipient_id", "recipient_name", "tokens_in", "tokens_out", "tokens_total", "tokens_cost"]
    report_line = [
        timestamp, account_id, account_name, entity_type, message, recipient_id, recipient_name,
        tokens["input"], tokens["output"], tokens["total"], tokens["cost"]
    ]

    if os.path.exists(REPORTS_FILE):
        workbook = openpyxl.load_workbook(REPORTS_FILE)
        sheet = workbook.active
    else:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        for col_idx, header in enumerate(headers, 1):
            sheet[f"{get_column_letter(col_idx)}1"] = header

    row = sheet.max_row + 1 if sheet.max_row > 1 else 2
    for col_idx, value in enumerate(report_line, 1):
        sheet[f"{get_column_letter(col_idx)}{row}"] = value

    workbook.save(REPORTS_FILE)

# Имитация печати
def simulate_typing(vk, peer_id, text_length):
    typing_duration = text_length // 3
    print(f"Имитирую печать на {typing_duration} секунд...")
    elapsed = 0
    while elapsed < typing_duration:
        vk.messages.setActivity(peer_id=peer_id, type='typing')
        time.sleep(5)
        elapsed += 5
        print(f"Имитирую печать... прошло {elapsed} секунд.")

# Очистка сообщения от символов **
def clean_message(message):
    return message.replace("**", "")

# Статус онлайн
def keep_online(vk):
    while True:
        try:
            vk.account.setOnline()
            print("Статус установлен на онлайн.")
        except Exception as e:
            print(f"Ошибка при установке статуса онлайн: {e}")
        time.sleep(300)

# Основная функция
def main():
    session_file = scan_sessions()
    if not session_file:
        print("Ошибка выбора или создания сессии.")
        return
    config = configparser.ConfigParser()
    config.read(session_file, encoding="utf-8")
    vk_token = config["DEFAULT"]["vk-token"]
    openai_token = config["DEFAULT"]["openai-token"]
    group_id = config["DEFAULT"].get("group-id", "")

    required_fields = {"tokens_in": "0", "tokens_out": "0", "tokens_total": "0", "tokens_cost": "0"}
    for field, default in required_fields.items():
        if field not in config["DEFAULT"]:
            config["DEFAULT"][field] = default
    with open(session_file, "w", encoding="utf-8") as configfile:
        config.write(configfile)

    client = OpenAI(api_key=openai_token)
    vk_session = vk_api.VkApi(token=vk_token)
    vk = vk_session.get_api()

    longpoll_params = {"access_token": vk_token, "v": "5.131"}
    if group_id:
        longpoll_params["group_id"] = group_id
    server_info = vk.messages.getLongPollServer(**longpoll_params)
    server = f"https://{server_info['server']}"
    key = server_info["key"]
    ts = server_info["ts"]

    if group_id:
        entity_id = f"-{group_id}"
        entity_name = config["DEFAULT"].get("group-name")
        if not entity_name:
            group_info = vk.groups.getById(group_id=group_id)[0]
            entity_name = group_info["name"]
            config["DEFAULT"]["group-name"] = entity_name
            with open(session_file, "w", encoding="utf-8") as configfile:
                config.write(configfile)
        print(f"LongPoll активирован для группы {entity_name} (ID: {group_id})")
    else:
        user_info = vk.users.get(user_ids=vk.users.get()[0]['id'], fields="first_name,last_name")[0]
        entity_id = user_info['id']
        entity_name = f"{user_info['first_name']} {user_info['last_name']}"
        print(f"LongPoll активирован для {entity_name} (ID: {entity_id})")

    online_thread = threading.Thread(target=keep_online, args=(vk,), daemon=True)
    online_thread.start()

    processing_peers = set()

    while True:
        try:
            response = retry_request(lambda timeout: requests.get(
                f"{server}?act=a_check&key={key}&ts={ts}&wait=25&mode=2&version=3", timeout=timeout))
            updates = response.json()

            if "failed" in updates:
                server_info = vk.messages.getLongPollServer(**longpoll_params)
                server = f"https://{server_info['server']}"
                key = server_info["key"]
                ts = server_info["ts"]
                continue

            ts = updates["ts"]
            for update in updates["updates"]:
                if update[0] != 4:
                    continue

                message_id = update[1]
                flags = update[2]
                peer_id = update[3]
                timestamp = update[4]
                text = update[5]
                extra_fields = update[6] if len(update) > 6 else {}

                user_id = extra_fields.get("from") if "from" in extra_fields else None
                if not user_id:
                    history_params = {"peer_id": peer_id, "count": 1}
                    if group_id:
                        history_params["group_id"] = group_id
                    last_message = vk.messages.getHistory(**history_params)["items"][0]
                    user_id = last_message["from_id"]

                if user_id == int(entity_id):
                    continue

                if group_id:
                    if peer_id < 2000000000:
                        sender_info = vk.users.get(user_ids=user_id, fields="first_name,last_name")[0]
                        if str(peer_id) != str(sender_info["id"]):
                            continue
                    elif peer_id >= 2000000000:
                        chat_info = vk.messages.getChat(chat_id=peer_id - 2000000000, fields="members")
                        members = [m["member_id"] for m in chat_info.get("members", [])]
                        if int(entity_id) not in members:
                            continue

                if peer_id in processing_peers:
                    print(f"Сообщение от {user_id} пропущено, так как уже обрабатывается предыдущее.")
                    continue

                processing_peers.add(peer_id)
                try:
                    sender_info = vk.users.get(user_ids=user_id, fields="first_name,last_name")[0]
                    print(f"Получено новое сообщение от {sender_info['first_name']} {sender_info['last_name']} (ID: {user_id}): {text}")

                    history_params = {"peer_id": peer_id, "count": 200}
                    if group_id:
                        history_params["group_id"] = group_id
                    messages = retry_request(lambda timeout: vk.messages.getHistory(**history_params, timeout=timeout))["items"]
                    if not messages:
                        print("Не удалось загрузить историю сообщений.")
                        continue

                    conversation_history = [{"from_id": msg["from_id"], "text": msg["text"], "date": msg["date"]} for msg in messages]
                    partner_info = get_conversation_partner_info(vk, user_id)
                    if not partner_info:
                        log_report(entity_id, entity_name, "", user_id, sender_info["first_name"] + " " + sender_info["last_name"], {"input": 0, "output": 0, "total": 0, "cost": 0})
                        continue
                    prompt = create_openai_prompt(config, conversation_history, partner_info, entity_id)
                    try:
                        response = client.chat.completions.create(model="gpt-4o-mini", messages=prompt, max_tokens=150)
                        reply = response.choices[0].message.content
                        input_tokens = response.usage.prompt_tokens
                        output_tokens = response.usage.completion_tokens
                        total_tokens = response.usage.total_tokens
                        tokens_cost = (input_tokens * 0.15 / 1000000) + (output_tokens * 0.6 / 1000000)

                        tokens_data = {
                            "input": input_tokens,
                            "output": output_tokens,
                            "total": total_tokens,
                            "cost": tokens_cost
                        }

                        simulate_typing(vk, peer_id, len(reply))
                        cleaned_reply = clean_message(reply)
                        # Используем retry_vk_request для отправки сообщения
                        retry_vk_request(lambda: vk.messages.send(
                            peer_id=peer_id,
                            message=cleaned_reply,
                            random_id=int(time.time() * 1000),
                            group_id=group_id if group_id else None
                        ))

                        update_session_tokens(session_file, tokens_data)
                        dossier_file = os.path.join(DOSSIER_DIR, f"{sender_info['first_name']}_{sender_info['last_name']}_{user_id}.json")
                        update_dossier_tokens(dossier_file, tokens_data)
                        log_report(entity_id, entity_name, reply, user_id, sender_info["first_name"] + " " + sender_info["last_name"], tokens_data)
                        print(f"Ответ от OpenAI: {reply}")
                    except Exception as e:
                        print(f"Ошибка отправки сообщения: {e}")
                finally:
                    processing_peers.remove(peer_id)
        except (ConnectionError, ReadTimeout, Exception) as e:
            print(f"Ошибка соединения или обработки: {e}")
            retries = 0
            backoff_factor = 1
            while True:
                retries += 1
                wait_time = min(backoff_factor * (2 ** retries), 60)
                print(f"Попытка переподключения. Ожидание {wait_time} секунд.")
                time.sleep(wait_time)
                try:
                    server_info = vk.messages.getLongPollServer(**longpoll_params)
                    server = f"https://{server_info['server']}"
                    key = server_info["key"]
                    ts = server_info["ts"]
                    print("Переподключение успешно.")
                    break
                except Exception as e:
                    print(f"Ошибка переподключения: {e}")

if __name__ == "__main__":
    main()
